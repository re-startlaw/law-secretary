#!/usr/bin/env python3
"""大中事件 類型証拠開示請求書別紙（甲1254-1258対応）の作成"""
import shutil
from copy import copy
from pathlib import Path
from openpyxl import load_workbook

SRC = Path("/Users/kometaninaoki/Library/CloudStorage/GoogleDrive-n.kometani@re-startlaw.com/.shortcut-targets-by-id/1pEJgbHuaj4L4ovFEISOuq-0LeHGInxYl/大中 忠生 : 詐欺未遂/ワードファイル/250918 類型証拠別紙案_宮村.xlsx")
DST = SRC.parent / "260427 類型証拠別紙案.xlsx"

shutil.copyfile(SRC, DST)

wb = load_workbook(DST)
ws = wb["別紙"]

# 既存データ行（row2-39）の書式を保持しつつ、値をクリア
old_max_row = ws.max_row
sample_styles = {}
for col_letter in ["A", "B", "C", "D", "E"]:
    c = ws[f"{col_letter}2"]
    sample_styles[col_letter] = {
        "font": copy(c.font),
        "fill": copy(c.fill),
        "alignment": copy(c.alignment),
        "border": copy(c.border),
        "number_format": c.number_format,
        "protection": copy(c.protection),
    }

# 古いデータ行を全部クリア（値も書式も）
for row in range(2, old_max_row + 1):
    for col_letter in ["A", "B", "C", "D", "E"]:
        cell = ws[f"{col_letter}{row}"]
        cell.value = None

REASON_FMT = "{kogo}の証明力を判断するためには、{noun}全ての開示を受けて、内容を比較検討することが重要であり、防御準備のためにその開示を受ける必要性は高い。"

companies = [
    ("甲1254", "レント株式会社"),
    ("甲1255", "デザインラボ株式会社"),
    ("甲1256", "TACコンサルタント株式会社"),
    ("甲1257", "京阪事務機器株式会社"),
    ("甲1258", "株式会社野村"),
]

rows_data = []

# 1行目: 内野大悟氏（1項5号ロ）
kogo1 = "甲1254-甲1258"
noun1 = "内野大悟氏の供述録取書等"
rows_data.append({
    "A": 1,
    "B": "1項5号ロ",
    "C": kogo1,
    "D": noun1,
    "E": REASON_FMT.format(kogo=kogo1, noun=noun1),
})

# 2-6行目: 各社の構成員（1項6号）
for i, (kogo, company) in enumerate(companies, start=2):
    noun = f"令和２年４月１日以降、{company}の株主、役員、従業員を含む構成員だったことがある者の供述録取書等"
    rows_data.append({
        "A": i,
        "B": "1項６号",
        "C": kogo,
        "D": noun,
        "E": REASON_FMT.format(kogo=kogo, noun=noun),
    })

# データ書き込み（書式も復元）
for idx, data in enumerate(rows_data, start=2):
    for col_letter in ["A", "B", "C", "D", "E"]:
        cell = ws[f"{col_letter}{idx}"]
        cell.value = data[col_letter]
        st = sample_styles[col_letter]
        cell.font = copy(st["font"])
        cell.fill = copy(st["fill"])
        cell.alignment = copy(st["alignment"])
        cell.border = copy(st["border"])
        cell.number_format = st["number_format"]
        cell.protection = copy(st["protection"])

wb.save(DST)
print(f"OK: {DST}")

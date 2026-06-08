#!/usr/bin/env python3
from pathlib import Path
from openpyxl import load_workbook

DST = Path("/Users/kometaninaoki/Library/CloudStorage/GoogleDrive-n.kometani@re-startlaw.com/.shortcut-targets-by-id/1pEJgbHuaj4L4ovFEISOuq-0LeHGInxYl/大中 忠生 : 詐欺未遂/ワードファイル/260427 類型証拠別紙案.xlsx")
wb = load_workbook(DST)
ws = wb["別紙"]
print(f"Max row: {ws.max_row}")
for row in ws.iter_rows(min_row=1, max_row=8, values_only=False):
    for cell in row:
        if cell.value is not None:
            print(f"  {cell.coordinate}: {cell.value}")
    print("---")

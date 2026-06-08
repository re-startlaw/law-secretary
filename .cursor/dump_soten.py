import openpyxl
from openpyxl.utils import get_column_letter

paths = [
"/Users/kometaninaoki/Library/CloudStorage/GoogleDrive-n.kometani@re-startlaw.com/マイドライブ/共有用/01_事件記録/ふ_メディアクリエイト・古澤智一/06資料/260617期日用/260605_争点整理表案（甲17のみ）_被告反論.xlsx",
"/Users/kometaninaoki/Library/CloudStorage/GoogleDrive-n.kometani@re-startlaw.com/マイドライブ/共有用/01_事件記録/ふ_メディアクリエイト・古澤智一/06資料/260617期日用/260605_争点整理表案（甲31のみ）_被告反論.xlsx",
]

for p in paths:
    print("#"*120)
    print("FILE:", p.split("/")[-1])
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb.active
    for r in range(1, ws.max_row+1):
        cells=[]
        for c in range(2, 20):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip()!="":
                cells.append(f"{get_column_letter(c)}{r}={v!r}")
        if cells:
            print(" | ".join(cells))

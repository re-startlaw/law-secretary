import openpyxl, sys

paths = [
"/Users/kometaninaoki/Library/CloudStorage/GoogleDrive-n.kometani@re-startlaw.com/マイドライブ/共有用/01_事件記録/ふ_メディアクリエイト・古澤智一/06資料/260617期日用/260605_争点整理表案（甲17のみ）_被告反論.xlsx",
"/Users/kometaninaoki/Library/CloudStorage/GoogleDrive-n.kometani@re-startlaw.com/マイドライブ/共有用/01_事件記録/ふ_メディアクリエイト・古澤智一/06資料/260617期日用/260605_争点整理表案（甲31のみ）_被告反論.xlsx",
]

for p in paths:
    print("="*100)
    print("FILE:", p.split("/")[-1])
    wb = openpyxl.load_workbook(p, data_only=True)
    for ws in wb.worksheets:
        print("--- SHEET:", ws.title, "dims:", ws.dimensions, "maxcol:", ws.max_column, "maxrow:", ws.max_row)

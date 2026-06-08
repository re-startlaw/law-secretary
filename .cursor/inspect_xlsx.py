#!/usr/bin/env python3
"""雛形Excelの内容を確認するスクリプト"""
import sys
from pathlib import Path
from openpyxl import load_workbook

SRC = Path("/Users/kometaninaoki/Library/CloudStorage/GoogleDrive-n.kometani@re-startlaw.com/.shortcut-targets-by-id/1pEJgbHuaj4L4ovFEISOuq-0LeHGInxYl/大中 忠生 : 詐欺未遂/ワードファイル/250918 類型証拠別紙案_宮村.xlsx")

wb = load_workbook(SRC, data_only=False)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print(f"Merged cells: {list(ws.merged_cells.ranges)}")
    print()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if cell.value is not None:
                print(f"  {cell.coordinate}: {repr(cell.value)}")
        print("---")

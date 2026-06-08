"""5原告の入学予定時期・不許可判明時期・経過日数まとめ。"""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


OUT = Path(
    "/Users/kometaninaoki/Library/CloudStorage/"
    "GoogleDrive-n.kometani@re-startlaw.com/マイドライブ/共有用/"
    "01_事件記録/い_インバウンドサポート学費訴訟/wordファイル/"
    "260518_5原告_入学予定_不許可判明_経過日数.xlsx"
)


rows = [
    {
        "no": 1,
        "name": "AHASAN（アハサン）",
        "enroll_date": date(2022, 10, 1),
        "enroll_label": "令和4年10月（2022年10月）",
        "denial_date": date(2022, 11, 8),
        "denial_label": "令和4年11月8日（2022年11月8日）",
        "note": "訴状では令和5年4月入学予定、被告通知は令和5年6月7日と記載。"
        "原告第7準備書面で令和4年10月入学・令和4年11月5日判明に訂正、"
        "原告第8準備書面で判明日を令和4年11月8日へ再訂正。",
    },
    {
        "no": 2,
        "name": "JIBON（ジボン）",
        "enroll_date": date(2023, 4, 1),
        "enroll_label": "令和5年4月（2023年4月）",
        "denial_date": date(2023, 5, 24),
        "denial_label": "令和5年5月24日（2023年5月24日）",
        "note": "訴状では被告通知日（令和5年6月7日）を判明日として記載。"
        "原告第7準備書面で実際の判明日を令和5年5月24日と明記。",
    },
    {
        "no": 3,
        "name": "MIAH（ミア）",
        "enroll_date": date(2023, 4, 1),
        "enroll_label": "令和5年4月（2023年4月）",
        "denial_date": date(2023, 4, 20),
        "denial_label": "令和5年4月20日（2023年4月20日）",
        "note": "訴状では被告通知日（令和5年10月3日）を判明日として記載。"
        "原告第7準備書面で令和4年4月3日と記載 → 原告第8準備書面で令和5年4月20日に訂正。",
    },
    {
        "no": 4,
        "name": "HOSAN（ホサン）",
        "enroll_date": date(2023, 4, 1),
        "enroll_label": "令和5年4月（2023年4月）",
        "denial_date": date(2023, 5, 24),
        "denial_label": "令和5年5月24日（2023年5月24日）",
        "note": "原告第7準備書面で令和5年5月24日と明記。以後訂正なし。",
    },
    {
        "no": 5,
        "name": "BANIK（バニク）",
        "enroll_date": date(2024, 4, 1),
        "enroll_label": "令和6年4月（2024年4月）",
        "denial_date": date(2024, 6, 3),
        "denial_label": "令和6年6月3日（2024年6月3日）",
        "note": "1回目は令和5年7月入学予定で在留資格認定証明書を取得済みだったが、"
        "パスポート有効期限の問題でビザ申請不能。"
        "パスポート再発行後、2回目の手続で令和6年4月入学を再申請。"
        "原告第7準備書面で判明日を令和6年6月3日と明記。以後訂正なし。",
    },
]


thin = Side(border_style="thin", color="808080")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF", size=11)
cell_font = Font(size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)


wb = Workbook()
ws = wb.active
ws.title = "5原告_経過日数"

headers = [
    "No.",
    "原告名",
    "入学予定時期",
    "ビザ不許可判明時期",
    "経過日数（入学予定→判明）",
    "備考（訂正経過等）",
]
ws.append(headers)
for col_idx, _ in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

for r in rows:
    delta_days = (r["denial_date"] - r["enroll_date"]).days
    ws.append([
        r["no"],
        r["name"],
        r["enroll_label"],
        r["denial_label"],
        f"{delta_days}日",
        r["note"],
    ])

for row_idx in range(2, 2 + len(rows)):
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = cell_font
        cell.border = border
        if col_idx in (1, 5):
            cell.alignment = center
        elif col_idx == 6:
            cell.alignment = left_wrap
        else:
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )

widths = [5, 22, 26, 28, 22, 70]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w

ws.row_dimensions[1].height = 28
for row_idx in range(2, 2 + len(rows)):
    ws.row_dimensions[row_idx].height = 90

note_row = 2 + len(rows) + 1
ws.cell(row=note_row, column=1, value="算定基準：")
ws.cell(
    row=note_row,
    column=2,
    value=(
        "入学予定時期が「○年○月」と月単位の場合、当該月の1日を起算日とした"
        "（米谷弁護士指示）。経過日数は『入学予定日 → 不許可判明日』の日数。"
    ),
).alignment = left_wrap
ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=6)
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=note_row, column=col_idx)
    cell.font = Font(italic=True, color="595959", size=10)

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"saved: {OUT}")
